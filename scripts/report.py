"""
[report.py] 엑셀 리포트 생성 모듈
──────────────────────────────────────────────────────
역할:
  - Top100 분석 결과를 6개 시트 엑셀 파일로 생성
  - 저장 위치: data/report_YYYYMMDD_YYYYMMDD.xlsx

시트 구성:
  1. 통합_TOP100   — 코스피+코스닥 전체 수익률 Top100 (수익률 순)
  2. 코스피_TOP100 — 코스피 단독 Top100 (수익률 순)
  3. 코스닥_TOP100 — 코스닥 단독 Top100 (수익률 순)
  4. 복합점수_TOP100 — 수익률60+ROE25+영업이익률15 가중합산 순위
  5. 재진입_포착   — 이전 기간 51~100위 → 이번 Top100 재진입 종목
  6. 관심종목      — data/watchlist.json 등록 종목 (없으면 시트 생략)

주요 함수:
  create_excel_report(combined, kospi, kosdaq, reentry, start, end,
                      watchlist_df=None)
      → 생성된 엑셀 파일 경로(str) 반환

컬럼 순서:
  수익률 시트: 종목코드 | 종목명 | 시작가 | 종료가 | 수익률(%) | ROE | 영업이익률
  복합점수 시트: 종목코드 | 종목명 | 시장 | 수익률(%) | ROE | 영업이익률 | 종합점수

의존성: pandas, openpyxl (pip install pandas openpyxl)

수정 이력:
  2026-03-19  최초 작성 — Phase 1-4
              stock-scanner/main.py create_excel_report() 기반 재구성
              헤더 스타일(굵게, 배경색) 및 컬럼 너비 자동 조정 추가
  2026-03-23  복합점수_TOP100 시트 추가 (수익률60+ROE25+영업이익률15 가중합산)
              엑셀 시트 수 5→6개
──────────────────────────────────────────────────────
"""

import os
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

# 수익률 시트 컬럼
DISPLAY_COLS = ["종목코드", "종목명", "시작가", "종료가", "수익률(%)",
                "ROE", "영업이익률"]

# 복합점수 시트 컬럼 (시장 구분 포함, 종합점수 강조)
SCORE_COLS = ["종목코드", "종목명", "시장", "수익률(%)", "ROE", "영업이익률", "종합점수"]

# 복합점수 시트 헤더 색상 (보라 계열로 구분)
SCORE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="5B4DB5")

# 헤더 스타일
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="2E5FA3")  # 파란 계열
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _style_sheet(ws) -> None:
    """워크시트 헤더 스타일 + 컬럼 너비 자동 조정

    Args:
        ws: openpyxl Worksheet 객체
    """
    # 헤더 행 스타일 적용
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # 컬럼 너비: 데이터 최대 길이 기준 자동 조정
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        # 한글 1자 ≈ 2자 너비로 보정, 최소 8 / 최대 30
        adjusted = min(max(max_len * 1.5, 8), 30)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _write_score_sheet(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    """복합점수_TOP100 시트 작성 — 보라색 헤더 + 종합점수 강조

    Args:
        writer: pd.ExcelWriter 인스턴스
        df:     복합점수 순 정렬된 통합 DataFrame
    """
    out_cols = [c for c in SCORE_COLS if c in df.columns]
    out_df = df[out_cols].copy()

    # 순위 컬럼 맨 앞에 추가
    out_df.insert(0, "순위", range(1, len(out_df) + 1))
    out_df.to_excel(writer, sheet_name="복합점수_TOP100", index=False)

    ws = writer.sheets["복합점수_TOP100"]

    # 헤더 스타일 (보라 계열)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = SCORE_HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # 종합점수 컬럼 강조 (마지막 컬럼)
    score_col_idx = out_df.columns.get_loc("종합점수") + 1 if "종합점수" in out_df.columns else None
    if score_col_idx:
        score_fill = PatternFill(fill_type="solid", fgColor="EDE9FF")  # 연보라 배경
        for row in ws.iter_rows(min_row=2, min_col=score_col_idx, max_col=score_col_idx):
            for cell in row:
                cell.fill = score_fill
                cell.font = Font(bold=True, color="3B1FA3")

    # 컬럼 너비 자동 조정
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len * 1.5, 8), 30)


def _write_sheet(writer: pd.ExcelWriter,
                 df: pd.DataFrame,
                 sheet_name: str,
                 cols: list[str] = None) -> None:
    """DataFrame을 지정 시트에 쓰고 스타일 적용

    Args:
        writer:     pd.ExcelWriter 인스턴스
        df:         출력할 DataFrame
        sheet_name: 시트명
        cols:       출력 컬럼 목록 (None이면 df 전체)
    """
    if cols:
        # df에 없는 컬럼은 조용히 제외
        out_cols = [c for c in cols if c in df.columns]
        df = df[out_cols]

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    _style_sheet(writer.sheets[sheet_name])


def create_excel_report(combined_df: pd.DataFrame,
                        kospi_df: pd.DataFrame,
                        kosdaq_df: pd.DataFrame,
                        reentry_df: pd.DataFrame,
                        start_date: str,
                        end_date: str,
                        watchlist_df: Optional[pd.DataFrame] = None) -> str:
    """5개 시트 엑셀 리포트 생성

    Args:
        combined_df:  통합 Top100 DataFrame
        kospi_df:     코스피 Top100 DataFrame
        kosdaq_df:    코스닥 Top100 DataFrame
        reentry_df:   재진입 종목 DataFrame (빈 DataFrame 허용)
        start_date:   분석 시작일 (YYYYMMDD)
        end_date:     분석 종료일 (YYYYMMDD)
        watchlist_df: 관심종목 DataFrame (None이면 시트 생략)

    Returns:
        생성된 엑셀 파일 경로 (6개 시트)
    """
    os.makedirs(DATA_DIR, exist_ok=True)
    filename = f"report_{start_date}_{end_date}.xlsx"
    filepath = os.path.join(DATA_DIR, filename)

    try:
        # 엑셀은 수익률 순으로 출력 (복합점수 정렬과 무관하게 유지)
        def _by_return(df: pd.DataFrame) -> pd.DataFrame:
            if df.empty or "수익률(%)" not in df.columns:
                return df
            return df.sort_values("수익률(%)", ascending=False).reset_index(drop=True)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # 1. 통합_TOP100
            _write_sheet(writer, _by_return(combined_df), "통합_TOP100", DISPLAY_COLS)

            # 2. 코스피_TOP100
            _write_sheet(writer, _by_return(kospi_df), "코스피_TOP100", DISPLAY_COLS)

            # 3. 코스닥_TOP100
            _write_sheet(writer, _by_return(kosdaq_df), "코스닥_TOP100", DISPLAY_COLS)

            # 4. 복합점수_TOP100 (combined_df는 이미 종합점수 순 정렬)
            if not combined_df.empty and "종합점수" in combined_df.columns:
                _write_score_sheet(writer, combined_df)
            else:
                empty = pd.DataFrame({"메시지": ["복합점수 데이터 없음 (분석 재실행 필요)"]})
                _write_sheet(writer, empty, "복합점수_TOP100")

            # 6. 재진입_포착
            if reentry_df.empty:
                empty = pd.DataFrame({"메시지": ["해당 기간 재진입 종목 없음"]})
                _write_sheet(writer, empty, "재진입_포착")
            else:
                _write_sheet(writer, reentry_df, "재진입_포착")

            # 7. 관심종목 (watchlist_df가 전달된 경우만)
            if watchlist_df is not None:
                if watchlist_df.empty:
                    empty = pd.DataFrame(
                        {"메시지": ["등록된 관심종목이 없거나 해당 기간 데이터 없음"]})
                    _write_sheet(writer, empty, "관심종목")
                else:
                    _write_sheet(writer, watchlist_df, "관심종목", DISPLAY_COLS)

    except Exception as e:
        raise RuntimeError(f"엑셀 생성 실패: {e}")

    print(f"[엑셀 저장] {filepath}")
    return filepath
